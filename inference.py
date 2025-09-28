import os, torch, argparse, tqdm, random, numpy as np
import pandas as pd
import sys
from util.dataset.base import create_data
from util.noiser.noiser import create_noiser
from util.model.model import create_model
from util.evaluate import perturbation_eval, deep_analysis
from openpyxl import load_workbook
import pickle

def seed_everything(seed):
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)


def match_ckpt(ckpt):
    _ckpt = {}
    for k, v in ckpt.items():
        if 'module.' in k:
            k = k.replace('network.module.', 'network.')
        _ckpt[k] = v
    return _ckpt


def main(args):
    args.global_rank = int(os.environ.get('RANK', 0))
    args.local_rank = int(os.environ.get('LOCAL_RANK', 0))
    args.node = int(args.global_rank // args.gpus)
    seed_everything(0 + args.global_rank)

    device = torch.device(f'cuda')
    _, _, prior_loader = create_data(args.prior, 1, batch_size=args.batch_size, data=args.data)
    pretrained_gene_list = prior_loader.dataset.gene_names
    prior_loader.dataset.pretrained_gene_list = pretrained_gene_list

    noiser = create_noiser(args.noiser, args, device)
    forward_model = create_model(prior_loader.dataset, args.method, args, device, noiser, rank=0, direction='f')
    backward_model = create_model(prior_loader.dataset, args.method, args, device, noiser, rank=0, direction='b')

    forward_model.to(device)
    backward_model.to(device)

    ckpt = torch.load(args.ckpt, map_location='cpu')
    forward_model.load_state_dict(match_ckpt(ckpt['forward_model']), strict=True)
    backward_model.load_state_dict(match_ckpt(ckpt['backward_model']), strict=True)

    with torch.no_grad():
        forward_model.eval()
        backward_model.eval()
        direction = 'q'

        save_path = os.path.join('./inference/', args.exp_name)
        epoch_name = args.ckpt.split("/")[-1]
        epoch_idx = epoch_name.split("_")[0]
        if not os.path.exists(save_path):
            os.makedirs(save_path)

        aug_graph = None
        true_list, ctrl_list, pred_list, cond_list = [], [], [], []

        it = iter(prior_loader)
        for bi in tqdm.tqdm(range(len(prior_loader)), desc=f' in {direction} inference'):
            batch = next(it)
            x_data = batch["data"].to(device, non_blocking=True)
            x_prior = batch["prior"].to(device, non_blocking=True)
            cond = batch["cond"]
            for key in cond.keys():
                cond[key] = cond[key].to(device, non_blocking=True)

            if 'gene' in args.prior and aug_graph is None:
                aug_graph = {
                    'G_go': prior_loader.dataset.G_go.to(device),
                    'G_go_weight': prior_loader.dataset.G_go_weight.to(device),
                }

            with torch.autocast(device_type="cuda", dtype=torch.float32, enabled=False):
                if direction == 'p':
                    pred = forward_model.inference(x_data, x_data, cond=cond, aug_graph=aug_graph, sample=True)[0][-1]
                    true, ctrl = x_prior, x_data
                else:
                    pred = backward_model.inference(x_prior, x_prior, cond=cond, aug_graph=aug_graph, sample=True)[0][-1]
                    true, ctrl = x_data, x_prior

            true_list.append(true)
            ctrl_list.append(ctrl)
            pred_list.append(pred)
            cond_list.append(cond['pert'])

        true = torch.cat(true_list).cpu()
        ctrl = torch.cat(ctrl_list).cpu()
        pred = torch.cat(pred_list).cpu()

        if 'gene' in args.prior:
            extras = prior_loader.dataset.extras
            true_conds = torch.cat(cond_list).cpu()
            de_gene_idx_dict = None if extras is None else extras.get("rank_genes_groups_cov_all_idx_dict")
            ndde20_idx_dict = None if extras is None else extras.get("top_non_dropout_de_20")
        else:
            de_gene_idx_dict = None
            ndde20_idx_dict = None
            true_conds = None
        gene_names = prior_loader.dataset.gene_names

        if true_conds is not None:
            score_dict, scores = perturbation_eval(
                true, pred, ctrl, gene_names=gene_names,
                de_gene_idx_dict=de_gene_idx_dict,
                true_conds=true_conds,
                ndde20_idx_dict=ndde20_idx_dict,
                path_to_save=f'{save_path}_{epoch_idx}_perturbation.png')
        else:
            scores = perturbation_eval(
                true, pred, ctrl,
                gene_names=gene_names,
                de_gene_idx_dict=de_gene_idx_dict,
                true_conds=true_conds,
                ndde20_idx_dict=ndde20_idx_dict,
                path_to_save=f'./inference/{args.data}_test_perturbation.png')
        for key in scores.keys():
            print(key, ":", scores[key])
        epoch_record = pd.DataFrame([scores], index=[epoch_idx])

        csv_filename = os.path.join(save_path, f'{args.data}_test_evaluation.xlsx')
        sheet_name = direction
        if not os.path.exists(csv_filename):
            with pd.ExcelWriter(csv_filename, engine='openpyxl') as writer:
                epoch_record.to_excel(writer, sheet_name=sheet_name)
        else:
            with pd.ExcelWriter(csv_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                writer._book = load_workbook(csv_filename)
                if sheet_name in writer._book.sheetnames:
                    startrow = writer._book[sheet_name].max_row
                    epoch_record.to_excel(writer, sheet_name=sheet_name, startrow=startrow, header=False)
                else:
                    epoch_record.to_excel(writer, sheet_name=sheet_name)


def create_parser():
    argparser = argparse.ArgumentParser()

    argparser.add_argument('--gpus', type=int, default=1, help='number of gpus per node')
    argparser.add_argument('--num_sample', type=int, default=128, help='number of samples')
    argparser.add_argument('--batch_size', type=int, default=16, help='batch size')

    argparser.add_argument('--method', type=str, default='dsb', help='method')
    argparser.add_argument('--simplify', action='store_true', help='whether to use simplified DSB')
    argparser.add_argument('--reparam', type=str, default=None,
                           help='whether to use reparameterized DSB, "term" for TR-DSB, "flow" for FR-DSB')
    argparser.add_argument('--noiser', type=str, default='dsb',
                           help='noiser type, "flow" noiser for Flow Matching models, "dsb" noiser for DSB models')
    argparser.add_argument('--gamma_type', type=str, default='linear_1e-4_1e-3', help='gamma schedule for DSB')
    argparser.add_argument('--training_timesteps', type=int, default=50, help='training timesteps')
    argparser.add_argument('--inference_timesteps', type=int, default=50, help='inference timesteps')

    argparser.add_argument('--network', type=str, default='gene-transformers', help='network architecture to use')
    argparser.add_argument('--use_amp', action='store_true', default=True,
                           help='whether to use mixed-precision training')

    argparser.add_argument('--data', type=str, default='norman', help='dataset')
    argparser.add_argument('--prior', type=str, default='genepert_test', help='prior distribution')

    argparser.add_argument('--exp_name', type=str, default='try', help='name of experiment')
    argparser.add_argument('--ckpt', type=str, default=None, help='checkpoint to load')
    argparser.add_argument('--config', type=str, default='./config/model.yaml', help='config file to load')

    return argparser


if __name__ == '__main__':

    argparser = create_parser()
    args = argparser.parse_args()
    if 'dsb' in args.method:
        assert args.training_timesteps == args.inference_timesteps
    main(args)
